import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def generate_peer_report():
    print("\n--- DAY 20: PEER COMPARISON REPORT GENERATOR (FIXED) ---")
    base_dir = Path(__file__).resolve().parent.parent.parent
    db_path = base_dir / 'data' / 'nifty100.db'
    output_path = base_dir / 'output' / 'peer_comparison.xlsx'
    peer_file = base_dir / 'data' / 'supporting' / 'peer_groups.xlsx'
    
    # 1. Ensure peer file exists
    if not peer_file.exists():
        print("[!] peer_groups.xlsx not found. Please run Day 18 script first.")
        return
        
    peers_df = pd.read_excel(peer_file)
    
    conn = sqlite3.connect(db_path)
    
    # 2. Get latest ratios (Without joining the old SQL sectors table)
    print("Extracting ratio data and percentiles...")
    ratios_df = pd.read_sql_query("""
        SELECT *
        FROM financial_ratios 
        WHERE year = (SELECT MAX(year) FROM financial_ratios r2 WHERE financial_ratios.company_id = r2.company_id)
    """, conn)
    
    # 3. Get percentiles
    pct_df = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)
    conn.close()
    
    if pct_df.empty:
        print("[!] No percentiles found. Ensure Day 18 script ran successfully.")
        return

    # Pivot percentiles to wide format so they sit next to their metrics
    pct_pivot = pct_df.pivot(index='company_id', columns='metric', values='percentile_rank')
    pct_pivot = pct_pivot.add_suffix('_pct_rank').reset_index()
    
    # Merge master dataset using the EXCEL peer groups to guarantee 11 sheets
    master_df = pd.merge(ratios_df, peers_df, on='company_id', how='inner')
    master_df = pd.merge(master_df, pct_pivot, on='company_id', how='inner')
    
    # 4. Create Excel Workbook
    wb = Workbook()
    wb.remove(wb.active) # Remove the default empty sheet
    
    # Define color fills
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    bold_font = Font(bold=True)
    
    print("Generating 11 peer group sheets with conditional formatting...")
    
    # Group by the peer_group_name from the Excel file
    for sector, group in master_df.groupby('peer_group_name'):
        sheet_name = str(sector).replace('/', '_')[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Sort by ROE descending so the best company is at the top (Benchmark)
        group = group.sort_values(by='return_on_equity_pct', ascending=False)
        
        # Write DataFrame to sheet
        for r in dataframe_to_rows(group, index=False, header=True):
            ws.append(r)
            
        # Format Header
        header_row = ws[1]
        for cell in header_row:
            cell.font = bold_font
            
        # Find which column indexes hold the percentile ranks
        pct_cols = [idx for idx, col in enumerate(group.columns, 1) if '_pct_rank' in col]
        
        # Apply conditional formatting
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
            if row_idx == 2:
                # Row 2 is the Benchmark company (Highest ROE). Highlight entire row gold.
                for cell in row:
                    cell.fill = gold_fill
            else:
                # Color code specific percentile cells
                for col_idx in pct_cols:
                    cell = row[col_idx - 1]
                    if isinstance(cell.value, (int, float)):
                        if cell.value >= 75:
                            cell.fill = green_fill
                        elif cell.value <= 25:
                            cell.fill = red_fill
                        else:
                            cell.fill = yellow_fill
                            
        # Add summary median row at the bottom
        ws.append([]) # Empty spacer row
        median_row = ["MEDIAN"]
        for col_name in group.columns[1:]:
            if pd.api.types.is_numeric_dtype(group[col_name]):
                median_row.append(group[col_name].median())
            else:
                median_row.append(None)
        
        ws.append(median_row)
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        
    wb.save(output_path)
    print(f"[✓] Successfully generated {output_path} with {len(wb.sheetnames)} peer group sheets.")

if __name__ == "__main__":
    generate_peer_report()
